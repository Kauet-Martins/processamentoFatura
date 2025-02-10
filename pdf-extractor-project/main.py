import os
from dotenv import load_dotenv
from extractor import extract_data_from_pdf
from spreadsheet import carregar_planilha
import openpyxl

load_dotenv()

def processar_faturas(pasta_pdfs, caminho_planilha):
    """Processa faturas da pasta e atualiza a planilha."""
    print(f"🔍 Iniciando processamento das faturas na pasta: {pasta_pdfs}")
    wb, ws, indice_coluna_uc = carregar_planilha(caminho_planilha)
    print("📊 Planilha carregada com sucesso!")
    
    for nome_pdf in os.listdir(pasta_pdfs):
        if nome_pdf.endswith(".pdf"):
            pdf_path = os.path.join(pasta_pdfs, nome_pdf)
            print(f"📂 Processando arquivo: {nome_pdf}")
            unidade_consumidora, valor_total = extract_data_from_pdf(pdf_path)
            
            if unidade_consumidora:
                print(f" Dados extraídos - UC: {unidade_consumidora}, Valor: {valor_total}")
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
                    uc_planilha = str(row[indice_coluna_uc]).strip().lstrip("0")
                    print(f"🔎 Verificando UC {uc_planilha} na planilha...")
                    if uc_planilha == unidade_consumidora:
                        for col in range(indice_coluna_uc + 1, len(row)):
                            if row[col] is None or row[col] == "":
                                ws.cell(row=i, column=col + 1, value=valor_total)
                                print(f" Valor {valor_total} inserido na linha {i}, coluna {col + 1}")
                                break
                        break
    
    wb.save(caminho_planilha)
    print(" Processamento concluído e planilha salva!")

def main():
    pasta_pdfs = os.getenv("PASTA_PDFS", "./pdfs")
    caminho_planilha = os.getenv("PLANILHA", "./planilha.xlsx")
    print(" Iniciando execução do script principal...")
    processar_faturas(pasta_pdfs, caminho_planilha)
    print(" Execução finalizada com sucesso!")

if __name__ == "__main__":
    main()

